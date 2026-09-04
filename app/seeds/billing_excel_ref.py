"""Чтение эталонных Excel биллинга Аристон (порт из Billings)."""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

# Эталон август 2026 (Ariston billing 08.2026.xlsx из fixtures)
CANONICAL_AUGUST_2026_PRR = {
    "manual_m3": Decimal("2867.014"),
    "mechanized_m3": Decimal("1564.572"),
    "billing_total": Decimal("9899958.14"),
}

PRR_DEDUPE_FIELDS = (
    "operation_date",
    "waybill",
    "vehicle",
    "op_type",
    "volume",
    "in_manual",
    "in_mech",
    "out_manual",
    "out_mech",
)

NEW_LINE_MAP = {
    "1.1": "storage_area_fixed",
    "1.2": "storage_area_extra",
    "2": "manual_m3",
    "2.0": "manual_m3",
    "2.1": "extra_manual_m3",
    "3": "mechanized_m3",
    "3.0": "mechanized_m3",
    "4": "vehicle_docs",
    "4.0": "vehicle_docs",
    "4.1": "vehicle_docs",
    "4.2": "extra_vehicle_docs",
    "4.3": "elco_passports",
    "5": "repack_units",
    "5.0": "repack_units",
    "5.1": "repack_units",
    "5.2": "valve_gluing",
    "5.3": "flue_stickering",
    "6": "overtime_m3",
    "7": "inventory_hours",
    "8": "elco_drain_hours",
}


def read_billing_reference(
  path: Path, month: int,
) -> tuple[dict[str, dict], Decimal | None, bool]:
  """Лист Billing → {line_code: {qty, amount}}, итого, two_tier (июн+)."""
  line_map = NEW_LINE_MAP if month >= 6 else {
    "1": "storage_area",
    "2": "manual_m3",
    "3": "mechanized_m3",
    "4": "vehicle_docs",
    "5": "repack_units",
    "6": "overtime_m3",
    "7": "inventory_hours",
    "8": "elco_drain_hours",
  }
  wb = load_workbook(path, read_only=True, data_only=True)
  ws = wb["Billing"]
  lines: dict[str, dict] = {}
  total = None
  for row in ws.iter_rows(min_row=4, max_row=50, values_only=True):
    if not row:
      continue
    num = str(row[0]).strip() if row[0] is not None else ""
    code = line_map.get(num)
    if code:
      qty = Decimal(str(row[4])) if row[4] not in (None, "-", "") else Decimal("0")
      amt = Decimal(str(row[5])) if row[5] is not None else Decimal("0")
      prev = lines.get(code)
      if prev:
        qty += prev["qty"]
        amt += prev["amount"]
      lines[code] = {"qty": qty, "amount": amt, "num": num, "name": row[1]}
    if row[4] == "ИТОГО:" and row[5] is not None and total is None:
      total = Decimal(str(row[5]))
  wb.close()
  return lines, total, month >= 6


def _prr_dec(value) -> Decimal:
    if value is None or value == "" or value == "-":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def waybill_key(value: str | None) -> str:
    s = (value or "").strip()
    return s.split()[0] if s else ""


def prr_row_dedupe_key(row: dict) -> tuple[str, ...]:
    return tuple(str(row.get(field) or "").strip() for field in PRR_DEDUPE_FIELDS)


def dedupe_prr_rows(rows: list[dict]) -> tuple[list[dict], int]:
    """Убрать точные дубликаты строк ПРР (частая ошибка в выгрузках)."""
    seen: set[tuple[str, ...]] = set()
    out: list[dict] = []
    skipped = 0
    for row in rows:
        key = prr_row_dedupe_key(row)
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        out.append(row)
    return out, skipped


def summarize_prr_handling_m3(rows: list[dict]) -> tuple[Decimal, Decimal]:
    """Суммы ручной/мех. обработки из ПРР (слияние по дате+накладной, как при сиде)."""
    by_waybill_day: dict[tuple[str, str], dict[str, Decimal]] = {}

    for raw in rows:
        day = str(raw.get("operation_date") or "")[:10]
        wb = waybill_key(str(raw.get("waybill") or ""))
        in_manual = _prr_dec(raw.get("in_manual"))
        in_mech = _prr_dec(raw.get("in_mech"))
        out_manual = _prr_dec(raw.get("out_manual"))
        out_mech = _prr_dec(raw.get("out_mech"))
        vol = _prr_dec(raw.get("volume"))

        key = (day, wb) if wb else (day, f"__{len(by_waybill_day)}")
        if wb and key in by_waybill_day:
            vo = by_waybill_day[key]
            vo["in_manual"] += in_manual
            vo["out_manual"] += out_manual
            vo["in_mech"] += in_mech
            vo["out_mech"] += out_mech
            vo["vol"] += vol
        else:
            by_waybill_day[key] = {
                "in_manual": in_manual,
                "out_manual": out_manual,
                "in_mech": in_mech,
                "out_mech": out_mech,
                "vol": vol,
            }

    manual = Decimal("0")
    mechanized = Decimal("0")
    for vo in by_waybill_day.values():
        m = vo["in_manual"] + vo["out_manual"]
        me = vo["in_mech"] + vo["out_mech"]
        manual += m
        mechanized += me + max(vo["vol"] - m - me, Decimal("0"))
    return manual, mechanized


def assert_prr_matches_billing(
    path: Path,
    month: int,
    *,
    tolerance: Decimal = Decimal("0.02"),
) -> tuple[Decimal, Decimal]:
    """ПРР должен давать те же объёмы, что строки 2/3 на листе Billing."""
    rows, skipped = dedupe_prr_rows(read_prr_rows(path))
    prr_manual, prr_mech = summarize_prr_handling_m3(rows)
    lines, _, _ = read_billing_reference(path, month)
    ref_manual = lines.get("manual_m3", {}).get("qty")
    ref_mech = lines.get("mechanized_m3", {}).get("qty")
    if ref_manual is None or ref_mech is None:
        raise ValueError(f"В {path.name} нет строк manual_m3/mechanized_m3 на листе Billing")

    errors: list[str] = []
    if abs(prr_manual - ref_manual) > tolerance:
        errors.append(
            f"ручная обработка: ПРР {prr_manual} ≠ Billing {ref_manual}"
        )
    if abs(prr_mech - ref_mech) > tolerance:
        errors.append(
            f"мехобработка: ПРР {prr_mech} ≠ Billing {ref_mech}"
        )
    if errors:
        hint = ""
        if skipped:
            hint = f" (убрано {skipped} точных дубликатов ПРР, объёмы всё равно не сходятся)"
        raise ValueError(
            f"ПРР не сходится с биллингом в {path.name}: "
            + "; ".join(errors)
            + hint
        )
    return prr_manual, prr_mech


def assert_august_2026_canonical_billing(path: Path) -> None:
    """Отклонить ошибочные выгрузки (дубли ПРР → 11468 / 6258 вместо эталона)."""
    lines, total, _ = read_billing_reference(path, month=8)
    canon = CANONICAL_AUGUST_2026_PRR
    errors: list[str] = []
    if total != canon["billing_total"]:
        errors.append(f"итого {total} ≠ эталон {canon['billing_total']}")
    manual = lines.get("manual_m3", {}).get("qty")
    mech = lines.get("mechanized_m3", {}).get("qty")
    if manual != canon["manual_m3"]:
        errors.append(f"ручная {manual} ≠ эталон {canon['manual_m3']}")
    if mech != canon["mechanized_m3"]:
        errors.append(f"мех. {mech} ≠ эталон {canon['mechanized_m3']}")
    if errors:
        raise ValueError(
            "Файл не является эталонным Ariston billing 08.2026.xlsx "
            f"(ожидается 2867.014 / 1564.572 м³ ПРР). " + "; ".join(errors)
        )


def read_prr_rows(path: Path) -> list[dict]:
  """Строки листа ПРР для импорта vehicle_operations."""
  wb = load_workbook(path, read_only=True, data_only=True)
  sheet = "ПРР" if "ПРР" in wb.sheetnames else wb.worksheets[1].title
  ws = wb[sheet]
  rows_iter = ws.iter_rows(values_only=True)
  header = next(rows_iter)
  col = {str(v).strip(): i for i, v in enumerate(header) if v}

  def idx(*names: str) -> int | None:
    for key, i in col.items():
      low = key.lower()
      if all(n.lower() in low for n in names):
        return i
    return None

  ix = {
    "operation_date": col.get("Дата отчета"),
    "waybill": col.get("Номер накладной"),
    "vehicle": col.get("Номер ТС:") or col.get("Номер ТС"),
    "op_type": col.get("Тип операции"),
    "volume": idx("объем", "документ") or col.get("Объем груза по документам (м³)"),
    "in_mech": idx("входящая", "механ") or col.get("Входящая поставка, механическая выгрузка, (м³)"),
    "in_manual": idx("входящая", "ручн") or col.get("Входящая поставка, ручная выгрузка, (м³)"),
    "out_mech": idx("исходящая", "механ") or col.get("Исходящая поставка, механизированная погрузка, (м³)"),
    "out_manual": idx("исходящая", "ручн") or col.get("Исходящая поставка, ручная погрузка, (м³)"),
    "reg_time": col.get("Время регистрации ТС в офисе:") or idx("регистрации", "тс"),
    "dep_time": col.get("Время убытия ТС:") or idx("убытия", "тс"),
    "torg2": col.get("№ акта ТОРГ-2"),
    "mx1": col.get("№ МХ-1"),
    "mx3": col.get("№ МХ-3"),
    "seal": col.get("№ пломбы"),
  }
  result: list[dict] = []
  for row in rows_iter:
    if not row or ix["operation_date"] is None or not row[ix["operation_date"]]:
      continue
    result.append({k: row[v] if v is not None else None for k, v in ix.items()})
  wb.close()
  return result
