"""Экспорт биллинга в Excel (формат для согласования с клиентом)."""
from __future__ import annotations

import calendar
import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.db import db
from app.modules.billing.calculator import BillingCalculator, load_contract_dict
from app.modules.reference.models import Client, Contract, ProductType
from app.modules.uss.models import VehicleOperation
from app.modules.uss.services.overtime import row_is_overtime

LINE_NUMBERS = {
    "storage_area_fixed": "1.1",
    "storage_area_extra": "1.2",
    "storage_area": "1",
    "manual_m3": "2.0",
    "mechanized_m3": "3.0",
    "vehicle_docs": "4.0",
    "repack_units": "5",
    "overtime_m3": "6",
    "inventory_hours": "7",
    "elco_drain_hours": "8",
    "warehouse_rent": "1",
    "office_rent": "2",
    "opex": "3",
}

PRR_HEADERS = [
    "Дата отчета",
    "Номер накладной",
    "Номер ТС:",
    "Тип операции",
    "Сверхурочные ПРР",
    "Время регистрации ТС в офисе:",
    "Время начала ПРР:",
    "Время окончания ПРР:",
    "Время убытия ТС:",
    "Объем груза по документам (м³)",
    "Входящая поставка, механическая выгрузка, (м³)",
    "Входящая поставка, ручная выгрузка, (м³)",
    "Исходящая поставка, механизированная погрузка, (м³)",
    "Исходящая поставка, ручная погрузка, (м³)",
]


def _month_bounds(year: int, month: int) -> tuple[date, date]:
    last = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last)


def _fmt_time(value) -> str | None:
    if value is None:
        return None
    return str(value)[:19]


def _op_type_label(code: str) -> str:
    return "Приёмка" if code == "inbound" else "Отгрузка"


def _overtime_label(flag: bool) -> str:
    return "Да" if flag else "Нет"


def _line_comment(line: dict) -> str | None:
    if line.get("formula_comment"):
        return line["formula_comment"]
    details = line.get("details") or []
    for d in details:
        if d.get("source_type") == "formula_comment" and d.get("description"):
            return d["description"]
    parts = [d.get("description") for d in details if d.get("description")]
    return "; ".join(parts) if parts else None


def _autosize_columns(ws, max_col: int, max_row: int):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        width = 12
        for row in range(1, max_row + 1):
            val = ws.cell(row, col).value
            if val is not None:
                width = max(width, min(len(str(val)) + 2, 50))
        ws.column_dimensions[letter].width = width


def _write_billing_sheet(ws, year: int, month: int, contract: dict, lines: list[dict]):
    ws["B1"] = "Расчётный период:"
    ws["C1"] = f"{month:02d}.{year}"
    if contract.get("number"):
        ws["D1"] = "Договор"
        ws["E1"] = contract["number"]

    headers = ["№", "Параметр", "Тариф", "Количество дней", "Количество единиц", "Сумма:", "Комментарий"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(3, col, title)
        cell.font = Font(bold=True)

    start_row = 4
    total = Decimal("0")
    for i, line in enumerate(lines):
        row = start_row + i
        num = LINE_NUMBERS.get(line["line_code"], str(i + 1))
        days = line.get("days_count")
        ws.cell(row, 1, num)
        ws.cell(row, 2, line["name"])
        ws.cell(row, 3, float(line["tariff"]) if line.get("tariff") is not None else None)
        ws.cell(row, 4, days if days else "-")
        ws.cell(row, 5, float(line["quantity"]))
        amount = Decimal(str(line["amount_ex_vat"]))
        ws.cell(row, 6, float(amount))
        comment = _line_comment(line)
        if comment:
            ws.cell(row, 7, comment)
        total += amount

    total_row = start_row + len(lines) + 1
    ws.cell(total_row, 5, "ИТОГО:").font = Font(bold=True)
    ws.cell(total_row, 6, float(total)).font = Font(bold=True)
    _autosize_columns(ws, 7, total_row)


def _vehicle_plate(op: VehicleOperation) -> str | None:
    if op.plate_number:
        return op.plate_number
    parts = [p for p in (op.tractor_plate, op.trailer_plate) if p]
    return " / ".join(parts) if parts else None


def _operation_export_dict(op: VehicleOperation) -> dict:
    rq = op.report_quantities or {}
    return {
        "operation_date": op.operation_date,
        "waybill_number": op.waybill_number,
        "vehicle_number": _vehicle_plate(op),
        "operation_type_code": op.operation_type_code or "inbound",
        "is_overtime": row_is_overtime(
            op.operation_date,
            departed_at=op.departed_at,
            warehouse_id=op.warehouse_id,
        ),
        "registered_at": op.registered_at,
        "prr_started_at": rq.get("prr_started_at"),
        "prr_finished_at": rq.get("prr_finished_at"),
        "departed_at": op.departed_at,
        "volume_document_m3": op.volume_document_m3,
        "inbound_mech_m3": rq.get("inbound_mech_m3", 0),
        "inbound_manual_m3": rq.get("inbound_manual_m3", 0),
        "outbound_mech_m3": rq.get("outbound_mech_m3", 0),
        "outbound_manual_m3": rq.get("outbound_manual_m3", 0),
        "extra_document_set_qty": op.extra_document_set_qty,
    }


def _write_prr_sheet(ws, operations: list[dict]):
    for col, title in enumerate(PRR_HEADERS, start=1):
        ws.cell(1, col, title).font = Font(bold=True)

    for i, op in enumerate(operations, start=2):
        ws.cell(i, 1, op["operation_date"])
        ws.cell(i, 2, op.get("waybill_number"))
        ws.cell(i, 3, op.get("vehicle_number"))
        ws.cell(i, 4, _op_type_label(op.get("operation_type_code", "inbound")))
        ws.cell(i, 5, _overtime_label(op.get("is_overtime", False)))
        ws.cell(i, 6, _fmt_time(op.get("registered_at")))
        ws.cell(i, 7, _fmt_time(op.get("prr_started_at")))
        ws.cell(i, 8, _fmt_time(op.get("prr_finished_at")))
        ws.cell(i, 9, _fmt_time(op.get("departed_at")))
        ws.cell(i, 10, float(op.get("volume_document_m3") or 0))
        ws.cell(i, 11, float(op.get("inbound_mech_m3") or 0))
        ws.cell(i, 12, float(op.get("inbound_manual_m3") or 0))
        ws.cell(i, 13, float(op.get("outbound_mech_m3") or 0))
        ws.cell(i, 14, float(op.get("outbound_manual_m3") or 0))

    _autosize_columns(ws, len(PRR_HEADERS), max(len(operations) + 1, 2))


def _write_waybills_sheet(ws, operations: list[dict]):
    ws["A1"] = "Основной комплект"
    ws["B1"] = "Дополнительный комплекты"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    main_waybills: list[str] = []
    extra_count = 0
    for op in operations:
        wb_num = (op.get("waybill_number") or "").strip()
        if wb_num and wb_num not in main_waybills:
            main_waybills.append(wb_num)
        extra_count += int(op.get("extra_document_set_qty") or 0)

    ws["A2"] = len(main_waybills)
    ws["B2"] = extra_count
    for i, wb_num in enumerate(main_waybills, start=3):
        ws.cell(i, 1, wb_num)

    _autosize_columns(ws, 2, max(len(main_waybills) + 2, 3))


def _write_details_sheet(ws, year: int, month: int, daily_areas: list[dict], tariff: float | None):
    ws["B1"] = "Площадь на начало месяца"
    if daily_areas:
        ws["C1"] = float(daily_areas[0].get("area_m2") or 0)

    headers = [
        "День месяца",
        "Занимаемая площадь хранения , м²",
        "Входящая поставка, м³",
        "Исходящая поставка, м³",
        "Количество пакетов документов (машин)",
        "сумма за день",
    ]
    for col, title in enumerate(headers, start=1):
        ws.cell(3, col, title).font = Font(bold=True)

    days = calendar.monthrange(year, month)[1]
    area_by_day = {}
    for row in daily_areas:
        d = row.get("snapshot_date") or row.get("date")
        if hasattr(d, "day"):
            area_by_day[d.day] = float(row.get("area_m2") or 0)

    rate = tariff or 0
    for day in range(1, days + 1):
        row_num = day + 3
        ws.cell(row_num, 1, day)
        area = area_by_day.get(day, ws["C1"].value or 0)
        ws.cell(row_num, 2, area)
        ws.cell(row_num, 6, rate * area)

    _autosize_columns(ws, 6, days + 3)


def _daily_area_from_line_details(lines: list[dict]) -> list[dict]:
    for line in lines:
        if line["line_code"] not in ("storage_area", "storage_area_fixed", "storage_area_extra"):
            continue
        out = []
        for d in line.get("details") or []:
            day = d.get("date") or d.get("detail_date")
            qty = d.get("quantity")
            if day and qty is not None:
                out.append({"snapshot_date": day, "area_m2": qty})
        if out:
            return out
    return []


def _load_vehicle_operations(contract_id: int, period_start: date, period_end: date) -> list[dict]:
    rows = (
        VehicleOperation.query.filter(
            VehicleOperation.contract_id == contract_id,
            VehicleOperation.operation_date >= period_start,
            VehicleOperation.operation_date <= period_end,
        )
        .order_by(VehicleOperation.operation_date, VehicleOperation.id)
        .all()
    )
    return [_operation_export_dict(r) for r in rows]


def build_billing_workbook(contract_id: int, year: int, month: int) -> tuple[bytes, str]:
    """Расчёт on-the-fly и сборка xlsx (без таблицы billing_lines)."""
    period_start, period_end = _month_bounds(year, month)
    contract = load_contract_dict(contract_id)
    if not contract:
        raise ValueError("contract_not_found")

    calc = BillingCalculator()
    result = calc.calculate_period(contract_id, period_start, period_end)
    if result.get("status") != "ok":
        raise ValueError(result.get("error") or "calculation_failed")

    lines = result.get("lines") or []
    full_contract = db.session.get(Contract, contract_id)
    client_name = ""
    product_code = contract.get("product_type_code") or ""
    if full_contract and full_contract.client_id:
        client = db.session.get(Client, full_contract.client_id)
        client_name = client.name if client else ""
    if full_contract and full_contract.product_type_id:
        pt = db.session.get(ProductType, full_contract.product_type_id)
        product_code = pt.code if pt else product_code

    config = contract.get("billing_config") or {}
    is_two_tier = config.get("area_mode", "simple") == "two_tier"
    operations = _load_vehicle_operations(contract_id, period_start, period_end)

    wb = Workbook()
    billing_ws = wb.active
    billing_ws.title = "Billing"
    _write_billing_sheet(billing_ws, year, month, contract, lines)

    if product_code == "RESPONSIBLE_STORAGE" and is_two_tier:
        prr_ws = wb.create_sheet("ПРР")
        _write_prr_sheet(prr_ws, operations)
        extra_ws = wb.create_sheet("Допкомплекты")
        _write_waybills_sheet(extra_ws, operations)
    else:
        details_ws = wb.create_sheet("Details")
        area_tariff = None
        for line in lines:
            if line["line_code"] in ("storage_area", "storage_area_fixed"):
                area_tariff = float(line["tariff"]) if line.get("tariff") is not None else None
                break
        daily = _daily_area_from_line_details(lines)
        _write_details_sheet(details_ws, year, month, daily, area_tariff)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    client_slug = (client_name or "client").replace(" ", "_")
    filename = f"{client_slug}_billing_{month:02d}.{year}.xlsx"
    return buf.getvalue(), filename
