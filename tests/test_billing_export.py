"""Экспорт биллинга в Excel."""
from io import BytesIO

import pytest
from openpyxl import load_workbook


def test_billing_export_endpoint(auth_client, client):
    auth_client("admin@test.local", "admin")
    resp = client.get("/api/billing/export?contract_id=1&year=2026&month=8")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.content_type
    assert ".xlsx" in resp.headers.get("Content-Disposition", "")

    wb = load_workbook(BytesIO(resp.data))
    assert "Billing" in wb.sheetnames
    ws = wb["Billing"]
    assert ws["B1"].value == "Расчётный период:"
    assert ws["C1"].value == "08.2026"


def test_billing_export_missing_params(auth_client, client):
    auth_client("admin@test.local", "admin")
    resp = client.get("/api/billing/export?contract_id=1")
    assert resp.status_code == 400


def test_billing_export_forbidden_without_role(client):
    resp = client.get("/api/billing/export?contract_id=1&year=2026&month=8")
    assert resp.status_code in (401, 403)
